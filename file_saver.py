from collections.abc import Iterable
from dataclasses import asdict
import json
from pathlib import Path
from time import time
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from allocator import Allocator, JuLei, Schulung

from config import COLUMN_WIDTH, HIGHLIGHT_COLOR, SHEET_WITH_JULEI_DATA, SHEET_WITH_SCHULUNGS_DATA, SHEET_WITH_SCORE_DATA
from config import FIRST_INDEX_IN_XLSX, FROM_BW_STRING, INITIAL_SHEET
from config import get_cell_index, get_column_index, get_row_index

from config import ALLOCATION_COLOR, DENIED_WISH_COLOR
from config import JULEI_FROM_BW_COLOR, JULEI_NOT_FROM_BW_COLOR, SCHULUNG_HEADER_COLOR, WHITE
from config import wish_from_bw_color, wish_not_from_bw_color

def _create_xlsx(first_sheet_name: str) -> Workbook:
    xlsx_file = Workbook()
    if len(xlsx_file.worksheets) == 0:  # should always be False after creation
        xlsx_file.create_sheet(first_sheet_name)
    first_sheet = xlsx_file.worksheets[0]
    first_sheet.title = first_sheet_name
    first_sheet.append(["Feel free to write or delete something here!"])
    first_sheet.append(["This sheet is not read, overwritten or needed by the computer."])
    return xlsx_file

class _Data:
    @staticmethod
    def add(data_containers: list[JuLei] | list[Schulung], sheet_name: str, xlsx_file: Workbook):
        if sheet_name in xlsx_file.sheetnames:
            del xlsx_file[sheet_name]
        sheet = xlsx_file.create_sheet(sheet_name)
        keys = [k for k in asdict(data_containers[0]).keys()]
        sheet.append(keys)
        for data_container in data_containers:
            sheet.append([json.dumps(asdict(data_container)[k]) for k in keys])

    @staticmethod
    def add_scores(a: Allocator, sheet_name: str, xlsx_file: Workbook):
        if sheet_name in xlsx_file.sheetnames:
            del xlsx_file[sheet_name]
        sheet = xlsx_file.create_sheet(sheet_name)
        for schulung in a.schulungen:
            sheet.append([a.scores[schulung][j] for j in a.juleis])

class _Insights:
    @staticmethod
    def draw_initial_first_row(a: Allocator, sheet: Any):
        for julei in a.juleis:
            cell_index = get_column_index(julei, a.juleis, True) + str(FIRST_INDEX_IN_XLSX)
            sheet[cell_index].font = Font(bold=True, color=WHITE)
            if julei.from_bw:
               sheet[cell_index] = FROM_BW_STRING
               sheet[cell_index].fill = PatternFill(start_color=JULEI_FROM_BW_COLOR, fill_type="solid")
            else:
               sheet[cell_index] = " "
               sheet[cell_index].fill = PatternFill(start_color=JULEI_NOT_FROM_BW_COLOR, fill_type="solid")

    @staticmethod
    def draw_initial_first_column(a: Allocator, sheet: Any):
        column_index = get_column_letter(FIRST_INDEX_IN_XLSX)
        sheet.column_dimensions[column_index].width = COLUMN_WIDTH

        for schulung in a.schulungen:
            cell_index = column_index + get_row_index(schulung, a.schulungen, True)
            sheet[cell_index] = schulung.capacity
            sheet[cell_index].font = Font(bold=True, color=WHITE)
            sheet[cell_index].fill = PatternFill(start_color=SCHULUNG_HEADER_COLOR, fill_type="solid")
            sheet[cell_index].alignment = Alignment(horizontal='center', vertical='center')

    @staticmethod
    def draw_initial_column(a: Allocator, sheet: Any, julei: JuLei):
        column_index = get_column_index(julei, a.juleis, True)
        sheet.column_dimensions[column_index].width = COLUMN_WIDTH

        for priority, schulungs_id in enumerate(julei.wishes):
            schulung = a.schulungen_by_id[schulungs_id]
            cell_index = get_cell_index(schulung, julei, a.schulungen, a.juleis, True)
            sheet[cell_index] = priority + 1
            sheet[cell_index].font = Font(color=WHITE)
            sheet[cell_index].alignment = Alignment(horizontal='center', vertical='center')
            if julei.from_bw:
                sheet[cell_index].fill = PatternFill(start_color=wish_from_bw_color(priority), fill_type="solid")
            else:
                sheet[cell_index].fill = PatternFill(start_color=wish_not_from_bw_color(priority), fill_type="solid")

    @staticmethod
    def draw_initial(a: Allocator, xlsx_file: Workbook):
        if INITIAL_SHEET in xlsx_file.sheetnames:
            del xlsx_file[INITIAL_SHEET]
        sheet = xlsx_file.create_sheet(INITIAL_SHEET)
        _Insights.draw_initial_first_row(a, sheet)
        _Insights.draw_initial_first_column(a, sheet)
        for julei in a.juleis:
            _Insights.draw_initial_column(a, sheet, julei)

    @staticmethod
    def draw_current_state(a: Allocator, sheet: Any):
        for schulung, juleis in a.participants.items():
            for julei in juleis:
                cell_index = get_cell_index(schulung, julei, a.schulungen, a.juleis, True)
                sheet[cell_index].fill = PatternFill(start_color=ALLOCATION_COLOR, fill_type="solid")

    @staticmethod
    def highlight_overfull_schulung(a: Allocator, sheet: Any, updated_schulung: Schulung, interested_juleis: Iterable[JuLei]):
        for interested_julei in interested_juleis:
            cell_index = get_cell_index(updated_schulung, interested_julei, a.schulungen, a.juleis, True)
            sheet[cell_index].fill = PatternFill(start_color=HIGHLIGHT_COLOR, fill_type="solid")

    @staticmethod
    def log_process_steps(a: Allocator, xlsx_file: Workbook):
        sheet = xlsx_file[INITIAL_SHEET]
        start_time = time()
        for updated_schulung, accepted_juleis, rejected_julei in a.set_participants():
            if not(rejected_julei is None):
                sheet = xlsx_file.copy_worksheet(sheet)
                sheet.title = f"{time() - start_time:.6f}"
                _Insights.draw_current_state(a, sheet)
                _Insights.highlight_overfull_schulung(a, sheet, updated_schulung, accepted_juleis + [rejected_julei])

            sheet = xlsx_file.copy_worksheet(sheet)
            sheet.title = f"{time() - start_time:.6f}"
            _Insights.draw_current_state(a, sheet)
            if not(rejected_julei is None):
                cell_index = get_cell_index(updated_schulung, rejected_julei, a.schulungen, a.juleis, True)
                sheet[cell_index].fill = PatternFill(start_color=DENIED_WISH_COLOR, fill_type="solid")

def save_to_xlsx(a: Allocator, xlsx_file_path: Path, verbose: bool=True):
    xlsx_file = _create_xlsx(a.name)

    if verbose:
        _Insights.draw_initial(a, xlsx_file)
        _Insights.log_process_steps(a, xlsx_file)

    _Data.add(a.juleis, SHEET_WITH_JULEI_DATA, xlsx_file)
    _Data.add(a.schulungen, SHEET_WITH_SCHULUNGS_DATA, xlsx_file)
    _Data.add_scores(a, SHEET_WITH_SCORE_DATA, xlsx_file)

    xlsx_file.save(xlsx_file_path)
