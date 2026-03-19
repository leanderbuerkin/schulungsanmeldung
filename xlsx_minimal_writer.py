from collections.abc import Iterable
from dataclasses import fields
from os import makedirs
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from data import Data, JuLei, Schulung
from xlsx_config import FIRST_INDEX_IN_XLSX, JULEI_SHEET_NAME, SCHULUNGEN_SHEET_NAME, SCORES_SHEET_NAME
from xlsx_config import get_new_workbook, index_as_columns, index_as_rows

def _add_juleis(juleis: Iterable[JuLei], xlsx: Workbook):
    sheet: Worksheet = xlsx.create_sheet(JULEI_SHEET_NAME)
    keys = [field.name for field in fields(JuLei)]
    for row, key in enumerate(keys, FIRST_INDEX_IN_XLSX):
        row = str(row)
        sheet["A" + row] = key
        for julei, column in index_as_columns(juleis).items():
            sheet[column + row] = json.dumps(getattr(julei, key))

def _add_schulungen(schulungen: Iterable[Schulung], xlsx: Workbook):
    sheet: Worksheet = xlsx.create_sheet(SCHULUNGEN_SHEET_NAME)
    keys = [field.name for field in fields(Schulung)]
    for column, key in enumerate(keys, FIRST_INDEX_IN_XLSX):
        column = get_column_letter(column)
        sheet[column + "1"] = key

        for schulung, row in index_as_rows(schulungen).items():
            sheet[column + row] = getattr(schulung, key)

def _add_scores(schulungen: Iterable[Schulung], juleis: Iterable[JuLei],
                scores: dict[Schulung, dict[JuLei, int]], xlsx: Workbook):
    sheet: Worksheet = xlsx.create_sheet(SCORES_SHEET_NAME)
    for julei, column in index_as_columns(juleis).items():
        for schulung, row in index_as_rows(schulungen).items():
            sheet[column + row] = scores[schulung][julei]

def get_minimal_xlsx(data: Data, output_directory: Path | None=None) -> Workbook:
    xlsx = get_new_workbook(data.name)
    return add_minimal_xlsx(xlsx, data, output_directory)

def add_minimal_xlsx(xlsx: Workbook, data: Data, output_directory: Path | None=None) -> Workbook:
    _add_schulungen(data.schulungen, xlsx)
    _add_juleis(data.juleis, xlsx)
    _add_scores(data.schulungen, data.juleis, data.scores, xlsx)

    if output_directory:
        makedirs(output_directory, exist_ok=True)
        xlsx.save(output_directory/f"{data.name}.xlsx")
    return xlsx