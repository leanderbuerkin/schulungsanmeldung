from enum import StrEnum
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

OUTPUT_DIRECTORY_PATH = Path("output")


FIRST_INDEX = 1
HEADER_INDIZES_COUNT = 1
FIRST_CONTENT_INDEX = FIRST_INDEX + HEADER_INDIZES_COUNT
COLUMN_WIDTH = 2.5

class SheetNames(StrEnum):
    STATS = "stats"
    EVENTS = "events"
    SEEKERS = "seekers"
    RANKED_WISHES = "wishes"
    RANKINGS = "rankings-generated"
    ORDERED_WISHES = "ordered-wishes-generated"
    PARTICIPANTS = "participants-generated"


def _get_indizes(amount: int) -> list[int]:
    offset = FIRST_CONTENT_INDEX
    return list(range(offset, offset + amount))

def get_row_indizes(amount: int) -> list[str]:
    return [str(i) for i in _get_indizes(amount)]

def get_column_indizes(amount: int) -> list[str]:
    return [get_column_letter(i) for i in _get_indizes(amount)]

def get_new_workbook() -> Workbook:
    xlsx = Workbook()
    xlsx.worksheets[0].append(["Switch sheets by pressing STRG + (SHIFT) + TAB."])
    return xlsx
