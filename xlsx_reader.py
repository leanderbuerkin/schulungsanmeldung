
from collections import defaultdict
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from allocator import Event, Seeker
from random_data_generator import Stats
from xlsx_config import SheetNames, get_column_index, get_row_index


def read_xlsx(xlsx_path: Path) -> dict[int, dict[Seeker, list[Event]]]:
    xlsx = load_workbook(xlsx_path)

    rows = xlsx[SheetNames.EVENTS].iter_rows(values_only=True)
    field_names = [str(field_name) for field_name in next(rows)]
    events_as_dicts = [
        _get_as_dict(row, field_names)
        for row in rows
    ]
    events = [
        Event(**event_as_dict)
        for event_as_dict in events_as_dicts
        if event_as_dict
    ]

    columns = xlsx[SheetNames.SEEKERS].iter_cols(values_only=True)
    field_names = [str(field_name) for field_name in next(columns)]
    seekers_as_dicts = [
        _get_as_dict(column, field_names)
        for column in columns
    ]
    seekers = [
        Seeker(**seeker_as_dict)
        for seeker_as_dict in seekers_as_dicts
        if seeker_as_dict
    ]

    sheet = xlsx[SheetNames.RANKED_WISHES]
    ranked_wishes: dict[int, dict[Seeker, list[Event]]] = defaultdict(dict)
    for seeker in seekers:
        for event in events:
            rank = sheet[get_column_index(seeker) + get_row_index(event)].value
            if not rank is None:
                if seeker not in ranked_wishes[rank].keys():
                    ranked_wishes[rank][seeker] = list()
                ranked_wishes[rank][seeker].append(event)

    return ranked_wishes


def read_stats_from_xlsx(xlsx: Workbook) -> Stats | None:
    sheet = xlsx[SheetNames.STATS]
    columns = sheet.iter_cols(values_only=True)
    field_names = [str(field_name) for field_name in next(columns)]
    stats_as_dict = _get_as_dict(next(columns), field_names)
    if stats_as_dict:
        return Stats(**stats_as_dict)


def _get_as_dict(cells: tuple[Any, ...], field_names: Iterable[str]) -> dict[str, Any] | None:
    data: dict[str, Any] = dict()
    for index, field_name in enumerate(field_names):
        if cells[index] is None:
            return None
        data[field_name] = cells[index]
    return data
