
from collections.abc import Generator, Iterable
from functools import partial
from os import makedirs
from pathlib import Path
from time import gmtime, strftime, time

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from _1_complete_data import CompleteData, XLSXWriter
from _0_input_data import Schulung, JuLei
from xlsx_config import XLSX, Colors

class ParticipantsLists:
    """
    Which JuLei participates in which Schulung is stored in two ways:

    - Once the participants for each Schulung
    - Once the Schulung each JuLei participates in

    This way we do not loose the information which Schulung stays empty
    and which JuLei does not participate in any Schulung.
    """
    _participants: dict[Schulung, list[JuLei]]
    _allocations: dict[JuLei, Schulung | None]

    _unchecked_wishes: dict[JuLei, list[Schulung]]
    _scores: dict[Schulung, dict[JuLei, int]]

    _overcrowded_schulungen: set[Schulung]

    @property
    def schulungen(self) -> list[Schulung]:
        schulungen = list(self._participants.keys())
        schulungen.sort(key=lambda s: (s.id, s.capacity))  # optional
        return schulungen

    @property
    def juleis(self) -> list[JuLei]:
        juleis = list(self._allocations.keys())
        juleis.sort(key=lambda j: (not j.from_bw, max(j.wishes)))  # optional
        return juleis

    @property
    def remaining_steps_max(self) -> int:
        """
        Worst case:
        Each step only one wish gets checked and all need to get checked.
        """
        return sum((len(self._unchecked_wishes[j]) for j in self.juleis))

    @property
    def can_be_improved(self) -> bool:
        if len(self._overcrowded_schulungen) > 0:
            return True
        for julei, schulung in self._allocations.items():
            if schulung is None and len(self._unchecked_wishes[julei]) > 0:
                return True
        return False

    def __init__(self, data: CompleteData):
        self._participants = {s: list() for s in data.schulungen}
        self._allocations = {j: None for j in data.juleis}
        self._unchecked_wishes = data.get_wishes_of_juleis()
        self._scores = data.scores
        self._overcrowded_schulungen = set()

    def get_participants(self, schulung: Schulung) -> list[JuLei]:
        self._allocate_as_many_juleis_as_possible()
        return self._participants[schulung]

    def get_allocation(self, julei: JuLei) -> Schulung | None:
        self._allocate_as_many_juleis_as_possible()
        return self._allocations[julei]

    # public for logging purposes
    def assign_juleis_ignoring_schulungs_capacity(self):
        for julei, wishes in self._unchecked_wishes.items():
            if self._allocations[julei] is None and len(wishes) > 0:
                desired_schulung = self._unchecked_wishes[julei].pop(0)
                self._add_allocations([(desired_schulung, julei)])

    # public for logging purposes
    def enforce_schulungs_capacity(self):
        while len(self._overcrowded_schulungen) > 0:
            schulung = next(iter(self._overcrowded_schulungen))

            # Lower properties are only considered, if the ones before are equal.
            self._participants[schulung].sort(key=lambda julei:(
                julei.from_bw,
                self._scores[schulung][julei]
            ))

            rejected_juleis = self._participants[schulung][:schulung.capacity]
            self._remove_allocations(((schulung, j) for j in rejected_juleis))

    def get_unfinished_participants(self) -> dict[Schulung, list[JuLei]]:
        return self._participants

    def _allocate_as_many_juleis_as_possible(self):
        while self.can_be_improved:
            self.assign_juleis_ignoring_schulungs_capacity()
            self.enforce_schulungs_capacity()

    def _add_allocations(self, allocations: Iterable[tuple[Schulung, JuLei]]):
        for schulung, julei in allocations:
            self._participants[schulung].append(julei)
            self._allocations[julei] = schulung
            if len(self._participants[schulung]) > schulung.capacity:
                self._overcrowded_schulungen.add(schulung)

    def _remove_allocations(self, allocations: Iterable[tuple[Schulung, JuLei]]):
        for schulung, julei in allocations:
            self._participants[schulung].remove(julei)
            self._allocations[julei] = None
            if len(self._participants[schulung]) <= schulung.capacity:
                if schulung in self._overcrowded_schulungen:
                    self._overcrowded_schulungen.remove(schulung)

class XLSXDrawer:
    COLUMN_WIDTH = 2.5
    @staticmethod
    def save_as_xlsx(data: CompleteData, output_directory: Path):
        pls = ParticipantsLists(data)
        rows = XLSX.as_rows(pls.schulungen)
        columns = XLSX.as_columns(pls.juleis)
        xlsx = XLSX.get_new_workbook(data.name)

        start_time = time()
        initial_sheet: Worksheet = xlsx.create_sheet(f"{time() - start_time:.6f}")
        XLSXDrawer._draw_initial_sheet(rows, columns, initial_sheet, data.get_wishes_of_juleis())

        sheet_generator = XLSXDrawer._get_sheet_generator(xlsx, initial_sheet, start_time)
        color_cells = partial(XLSXDrawer._color_cells, rows, columns)

        steps_counter = 0
        while pls.can_be_improved:
            pls.assign_juleis_ignoring_schulungs_capacity()
            sheet = next(sheet_generator)
            for schulung, candidates in pls.get_unfinished_participants().items():
                color_cells(sheet, [schulung], candidates, Colors.HIGHLIGHTED_WISH)
            sheet = next(sheet_generator)
            for schulung, candidates in pls.get_unfinished_participants().items():
                color_cells(sheet, [schulung], candidates, Colors.REJECTED_WISH)

            pls.enforce_schulungs_capacity()
            for schulung, participants in pls.get_unfinished_participants().items():
                color_cells(sheet, [schulung], participants, Colors.GRANTED_WISH)

            steps_counter += 1
            Logger.print_stats(start_time, steps_counter, pls)
        print(f"Finished with {steps_counter} steps in {strftime("%H:%M:%S", gmtime(time()-start_time))}!")

        xlsx = XLSXWriter.add_minimal_data(xlsx, data)

        makedirs(output_directory, exist_ok=True)
        xlsx.save(output_directory/f"{data.name}.xlsx")

    @staticmethod
    def _draw_initial_sheet(
            rows: dict[Schulung, str],
            columns: dict[JuLei, str],
            initial_sheet: Worksheet,
            wishes: dict[JuLei, list[Schulung]]
        ):
        for column in ["A"] + list(columns.values()):
            initial_sheet.column_dimensions[column].width = XLSXDrawer.COLUMN_WIDTH
        XLSXDrawer._draw_headers(rows, columns, initial_sheet)
        XLSXDrawer._draw_wishes(rows, columns, initial_sheet, wishes)

    @staticmethod
    def _draw_headers(rows: dict[Schulung, str], columns: dict[JuLei, str], sheet: Worksheet):
        sheet["A1"].fill = PatternFill("solid", Colors.SCHULUNG_HEADER)

        for schulung, row in rows.items():
            sheet["A"+row] = schulung.capacity
            sheet["A"+row].font = Font(bold=True, color=Colors.WHITE)
            sheet["A"+row].fill = PatternFill("solid", Colors.SCHULUNG_HEADER)
            sheet["A"+row].alignment = Alignment(horizontal='center')

        for julei, column in columns.items():
            sheet[column+"1"].font = Font(bold=True, color=Colors.WHITE)
            if julei.from_bw:
               sheet[column+"1"] = XLSX.FROM_BW_STRING
               sheet[column+"1"].fill = PatternFill("solid", Colors.JULEI_FROM_BW_HEADER)
            else:
               sheet[column+"1"] = XLSX.NOT_FROM_BW_STRING
               sheet[column+"1"].fill = PatternFill("solid", Colors.JULEI_NOT_FROM_BW_HEADER)
    
    @staticmethod
    def _draw_wishes(
            rows: dict[Schulung, str],
            columns: dict[JuLei, str],
            sheet: Worksheet,
            wishes: dict[JuLei, list[Schulung]]
        ):
        for julei, column in columns.items():
            for priority, row in enumerate(rows[s] for s in wishes[julei]):
                sheet[column+row] = priority + 1
                sheet[column+row].font = Font(color=Colors.WHITE)
                sheet[column+row].alignment = Alignment(horizontal='center')
                sheet[column+row].fill = PatternFill("solid", Colors.wish(priority, julei.from_bw))

    @staticmethod
    def _color_cells(
            rows: dict[Schulung, str],
            columns: dict[JuLei, str],
            sheet: Worksheet,
            schulungen: Iterable[Schulung],
            juleis: list[JuLei],
            color: str
        ):
        for row in (rows[s] for s in schulungen):
            for column in (columns[j] for j in juleis):
                sheet[column+row].fill = PatternFill("solid", color)
    
    @staticmethod
    def _get_sheet_generator(
            xlsx: Workbook,
            sheet_for_duplication: Worksheet,
            start_time: float
        ) -> Generator[Worksheet]:
        sheet = sheet_for_duplication
        while True:
            sheet = xlsx.copy_worksheet(sheet)
            sheet.title = f"{time() - start_time:.6f}"
            yield sheet

class Logger:
    @staticmethod
    def print_stats(start_time: float, steps_counter: int, pls: ParticipantsLists):
        processing_time = time()-start_time
        total_steps_max = steps_counter + pls.remaining_steps_max
        progress = steps_counter/total_steps_max
        total_time = processing_time/progress
        print(
            f"----------------{strftime("%H:%M:%S", gmtime(processing_time))}-------------\n" +
            f"Finished Steps: {steps_counter} in \n" +
            f"Remaining time: {strftime("%H:%M:%S", gmtime(total_time - processing_time))} (worst case)"
        )
