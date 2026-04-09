from enum import StrEnum
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from allocator import Event, Seeker

OUTPUT_DIRECTORY_PATH = Path("output")


FIRST_INDEX = 1
HEADER_INDIZES_COUNT = 1
CELL_WIDTH = 2.5
CELL_HEIGHT = 5

class SheetNames(StrEnum):
    STATS = "stats"
    EVENTS = "events"
    SEEKERS = "seekers"
    RANKED_WISHES = "wishes"

def get_new_workbook() -> Workbook:
    xlsx = Workbook()
    xlsx.worksheets[0].append(["Switch sheets by pressing STRG + (SHIFT) + TAB."])
    return xlsx

def get_column_index(seeker: Seeker) -> str:
    return get_column_letter(seeker.index + FIRST_INDEX + HEADER_INDIZES_COUNT)

def get_row_index(event: Event) -> str:
    return str(event.index + FIRST_INDEX + HEADER_INDIZES_COUNT)
