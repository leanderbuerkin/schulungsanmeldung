from dataclasses import asdict, fields
from datetime import datetime
from os import makedirs

from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from allocator import Event, Seeker
from random_data_generator import Stats
from xlsx_config import FIRST_INDEX, OUTPUT_DIRECTORY_PATH, SheetNames, get_column_index, get_new_workbook, get_row_index

def save_to_xlsx(
        ranked_wishes: dict[int, dict[Seeker, list[Event]]],
        participants_lists: dict[Event, list[Seeker]],
        stats: Stats | None,
        name: str = datetime.now().strftime("%Y_%m_%d")
    ):
    events: set[Event] = set()
    seekers: set[Seeker] = set()
    for wishes in ranked_wishes.values():
        for seeker, wished_events in wishes.items():
            seekers.add(seeker)
            for event in wished_events:
                events.add(event)
    for event, participating_seekers in participants_lists.items():
        events.add(event)
        for seeker in participating_seekers:
            seekers.add(seeker)

    xlsx = get_new_workbook()

    if stats:
        sheet = xlsx.create_sheet(SheetNames.STATS)
        field_names = [field.name for field in fields(Stats)]
        for row_index, field_name in enumerate(field_names, FIRST_INDEX):
            row = str(row_index)
            sheet["A" + row] = field_name
            sheet["B" + row] = getattr(stats, field_name)

    sheet = xlsx.create_sheet(SheetNames.EVENTS)
    field_names = [field.name for field in fields(Event)]
    for column_index, field_name in enumerate(field_names, FIRST_INDEX):
        column = get_column_letter(column_index)
        sheet[column + "1"] = field_name
        for event in events:
            sheet[column + get_row_index(event)] = getattr(event, field_name)

    sheet = xlsx.create_sheet(SheetNames.SEEKERS)
    field_names = [field.name for field in fields(Seeker)]
    for row_index, field_name in enumerate(field_names, FIRST_INDEX):
        row = str(row_index)
        sheet["A" + row] = field_name
        for seeker in seekers:
            sheet[get_column_index(seeker) + row] = getattr(seeker, field_name)

    sheet = xlsx.create_sheet(SheetNames.RANKED_WISHES)
    for event in events:
        sheet["A" + get_row_index(event)] = str(event)
        sheet["A" + get_row_index(event)].alignment = Alignment(horizontal='fill')
    for seeker in seekers:
        sheet[get_column_index(seeker) + "1"] = str(seeker)
    for rank, wishes in ranked_wishes.items():
        for seeker, wished_events in wishes.items():
            for event in wished_events:
                sheet[get_column_index(seeker) + get_row_index(event)] = rank

    for event, participating_seekers in participants_lists.items():
        for seeker in participating_seekers:
            sheet[get_column_index(seeker) + get_row_index(event)].fill = PatternFill("solid", "AAFFAA")

    if stats:
        name += "".join(f"__{key}_{max(0, value)}" for key, value in asdict(stats).items())
    makedirs(OUTPUT_DIRECTORY_PATH, exist_ok=True)
    xlsx.save(OUTPUT_DIRECTORY_PATH/f"{name}.xlsx")
