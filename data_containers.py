from dataclasses import dataclass
from os import makedirs
from pathlib import Path
from random import shuffle

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config import FIRST_CONTENT_COLUMN_INDEX, FIRST_CONTENT_ROW_INDEX

type UniqueJuLeiId = int
type UniqueSchulungsId = int

@dataclass(frozen=True)
class JuLei:
    id: UniqueJuLeiId
    from_bw: bool
    wishes: tuple[UniqueSchulungsId, ...]

@dataclass(frozen=True)
class Schulung:
    id: UniqueSchulungsId
    capacity: int

class Problem:
    name: str
    output_directory: Path
    xlsx_file: Workbook
    from_bw: dict[JuLei, bool]
    remaining_wishes: dict[JuLei, list[Schulung]]
    capacity: dict[Schulung, int]
    scores: dict[Schulung, dict[JuLei, int]]
    participants: dict[Schulung, list[JuLei]]
    columns_in_xlsx: dict[JuLei, str]
    rows_in_xlsx: dict[Schulung, str]
    number_of_possible_allocations: int

    # All of these properties are mainly for logging and estimating processing duration
    @property
    def number_of_juleis(self) -> int:
        return len(self.from_bw)

    @property
    def number_of_schulungen(self) -> int:
        return len(self.participants)

    @property
    def total_capacity(self) -> int:
        return sum(self.capacity.values())

    @property
    def juleis_per_slot(self) -> float:
        return self.number_of_juleis/self.total_capacity

    @property
    def number_of_unchecked_allocations(self) -> int:
        return sum((len(ws) for ws in self.remaining_wishes.values()))

    @property
    def number_of_checked_allocations(self) -> int:
        return self.number_of_possible_allocations - self.number_of_unchecked_allocations

    @property
    def expected_final_number_of_checked_allocations(self) -> int:
        return int(min(100, self.juleis_per_slot) * self.number_of_possible_allocations)

    @property
    def worst_case_final_number_of_checked_allocations(self) -> int:
        return self.number_of_possible_allocations

    @staticmethod
    def _get_demand(juleis: list[JuLei], schulungen: list[Schulung]) -> dict[Schulung, int]:
        demand = {s: - s.capacity for s in schulungen}
        for julei in juleis:
            for schulungs_id in julei.wishes:
                schulung = next(filter(lambda s: s.id == schulungs_id, schulungen))
                demand[schulung] += 1
        return demand

    def __init__(self, name: str, output_directory: Path, juleis: list[JuLei], schulungen: list[Schulung]):
        self.name = name
        self.output_directory = output_directory
        makedirs(self.output_directory, exist_ok=True)
        self.xlsx_file = Workbook()


        juleis = sorted(juleis, key=lambda j: (not j.from_bw, len(j.wishes)))
        demand = Problem._get_demand(juleis, schulungen)
        schulungen = sorted(schulungen, key=lambda s: (demand[s], s.capacity))

        self.from_bw = {j: j.from_bw for j in juleis}

        self.remaining_wishes = {j: list() for j in juleis}
        for julei in juleis:
            for schulungs_id in julei.wishes:
                schulung = next(filter(lambda s: s.id == schulungs_id, schulungen))
                self.remaining_wishes[julei].append(schulung)
        self.capacity = {s: s.capacity for s in schulungen}

        unique_scores = list(range(len(juleis)))
        self.scores = {s: dict() for s in schulungen}
        for s in schulungen:
            shuffle(unique_scores)
            self.scores[s] = {j: v for j, v in zip(juleis, unique_scores)}

        self.participants = {s: list() for s in schulungen}

        self.columns_in_xlsx = {j: get_column_letter(FIRST_CONTENT_COLUMN_INDEX + i) for i, j in enumerate(juleis)}
        self.rows_in_xlsx = {s: str(FIRST_CONTENT_ROW_INDEX + i) for i, s in enumerate(schulungen)}
        self.number_of_possible_allocations = self.number_of_unchecked_allocations

    def get_allocation(self, julei: JuLei) -> Schulung | None:
        for schulung, allocated_juleis in self.participants.items():
            if julei in allocated_juleis:
                return schulung
        return None

    def is_full(self, schulung: Schulung) -> bool:
        return len(self.participants[schulung]) >= self.capacity[schulung]

    def save_to_file(self):
        xlsx_file_path = self.output_directory/f"{self.name}.xlsx"
        self.xlsx_file.save(xlsx_file_path)
