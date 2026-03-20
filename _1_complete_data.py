"""
It is called "complete data" because all other data
can be generated from this

Add this stage, the participants list is already definied,
but is still some work to get it.
"""
from collections import defaultdict
from collections.abc import Iterable
from dataclasses import dataclass, fields
import json
from os import makedirs
from pathlib import Path
from random import shuffle
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from xlsx_config import XLSX
from _0_input_data import InputData, Schulung, JuLei, UniqueSchulungsId, UniqueJuLeiId

@dataclass(frozen=True)
class CompleteData:
    """
    If two JuLeis compete for a slot of a Schulung,
    the JuLei with the higher score for this Schulung gets the slot.
    """
    name: str
    schulungen_by_id: dict[UniqueSchulungsId, Schulung]
    juleis_by_id: dict[UniqueJuLeiId, JuLei]
    scores: dict[Schulung, dict[JuLei, int]]

    @property
    def schulungen(self) -> list[Schulung]:
        schulungen = list(self.schulungen_by_id.values())
        schulungen.sort(key=lambda s: (s.id, s.capacity))  # optional
        return schulungen

    @property
    def juleis(self) -> list[JuLei]:
        juleis = list(self.juleis_by_id.values())
        juleis.sort(key=lambda j: (not j.from_bw, max(j.wishes)))  # optional
        return juleis

    def get_wishes_of_juleis(self) -> dict[JuLei, list[Schulung]]:
        wishes_of_juleis: dict[JuLei, list[Schulung]] = defaultdict(list)
        for julei in self.juleis:
            for schulungs_id in julei.wishes:
                wish_of_julei = self.schulungen_by_id[schulungs_id]
                wishes_of_juleis[julei].append(wish_of_julei)

        return wishes_of_juleis

class RandomGenerator:
    @staticmethod
    def get_complete_data_from_input_data(data: InputData) -> CompleteData:
        name = data.name
        schulungen_by_id = {s.id: s for s in data.schulungen}
        juleis_by_id = {j.id: j for j in data.juleis}
        scores = RandomGenerator._get_random_scores(
            schulungen_by_id.values(), list(juleis_by_id.values())
        )

        return CompleteData(name, schulungen_by_id, juleis_by_id, scores)

    @staticmethod
    def _get_random_scores(
            schulungen: Iterable[Schulung],
            juleis: list[JuLei]
        ) -> dict[Schulung, dict[JuLei, int]]:
        scores: dict[Schulung, dict[JuLei, int]] = defaultdict(dict)
        for s in schulungen:
            shuffle(juleis)
            for score, julei in enumerate(juleis):
                scores[s][julei] = score

        return scores

class XLSXReader:
    @staticmethod
    def read_from_xlsx(xlsx_path: Path) -> CompleteData:
        name = xlsx_path.stem
        xlsx = load_workbook(xlsx_path)

        schulungen = XLSXReader._get_schulungen(xlsx)
        juleis = XLSXReader._get_juleis(xlsx)

        scores: dict[Schulung, dict[JuLei, int]] = defaultdict(dict)
        for schulung, row in XLSX.as_rows(schulungen).items():
            for julei, column in XLSX.as_columns(juleis).items():
                score = xlsx[XLSX.SCORES_SHEET_NAME][column+row].value
                scores[schulung][julei] = score

        juleis_by_id = {j.id: j for j in juleis}
        schulungen_by_id = {s.id: s for s in schulungen}

        return CompleteData(name, schulungen_by_id, juleis_by_id, scores)

    @staticmethod
    def _get_schulungen(xlsx: Workbook) -> set[Schulung]:
        sheet = xlsx[XLSX.SCHULUNGEN_SHEET_NAME]
        rows = sheet.iter_rows(values_only=True)
        field_names = [str(k) for k in next(rows)]

        schulungen_as_dicts: list[dict[str, Any]] = list()
        for row in rows:
            schulung_as_dict = XLSXReader._get_as_dict(row, field_names)
            schulungen_as_dicts.append(schulung_as_dict)

        return {Schulung(**d) for d in schulungen_as_dicts}
    
    @staticmethod
    def _get_juleis(xlsx: Workbook) -> set[JuLei]:
        sheet = xlsx[XLSX.JULEI_SHEET_NAME]
        columns = sheet.iter_cols(values_only=True)
        field_names = [str(k) for k in next(columns)]

        juleis_as_dicts: list[dict[str, Any]] = list()
        for column in columns:
            julei_as_dict = XLSXReader._get_as_dict(column, field_names)
            juleis_as_dicts.append(julei_as_dict)

        return {JuLei(**d) for d in juleis_as_dicts}

    @staticmethod
    def _get_as_dict(cells: tuple[Any, ...], field_names: Iterable[str]) -> dict[str, Any]:

        as_dict: dict[str, Any] = dict()
        for i, field_name in enumerate(field_names):
            as_dict[field_name] = json.loads(str(cells[i]))

        return XLSXReader.convert_lists_back_to_tuples(as_dict)

    @staticmethod
    def convert_lists_back_to_tuples(
            dict_from_json: dict[str, Any]
        ) -> dict[str, Any]:
        """JSON does not know tuples and converts them to lists."""

        for key, value in dict_from_json.items():
            if isinstance(value, list):
                dict_from_json[key] = tuple(value) # pyright: ignore[reportUnknownArgumentType]

        return dict_from_json

class XLSXWriter:
    @staticmethod
    def save_as_xlsx(data: CompleteData, output_directory: Path):
        xlsx = XLSXWriter.get_minimal_xlsx(data)
        makedirs(output_directory, exist_ok=True)
        xlsx.save(output_directory/f"{data.name}.xlsx")

    @staticmethod
    def get_minimal_xlsx(data: CompleteData) -> Workbook:
        xlsx = XLSX.get_new_workbook(data.name)
        xlsx = XLSXWriter.add_minimal_data(xlsx, data)
        return xlsx

    @staticmethod
    def add_minimal_data(xlsx: Workbook, data: CompleteData) -> Workbook:
        XLSXWriter._add_schulungen(data.schulungen, xlsx)
        XLSXWriter._add_juleis(data.juleis, xlsx)
        XLSXWriter._add_scores(data.schulungen, data.juleis, data.scores, xlsx)
        return xlsx

    @staticmethod
    def _add_juleis(juleis: list[JuLei], xlsx: Workbook):
        sheet = xlsx.create_sheet(XLSX.JULEI_SHEET_NAME)
        field_names = [field.name for field in fields(JuLei)]

        for row_index, field_name in enumerate(field_names, XLSX.FIRST_INDEX):
            row = str(row_index)
            sheet["A" + row] = field_name

            for julei, column in XLSX.as_columns(juleis).items():
                sheet[column+row] = json.dumps(getattr(julei, field_name))

    @staticmethod
    def _add_schulungen(schulungen: list[Schulung], xlsx: Workbook):
        sheet = xlsx.create_sheet(XLSX.SCHULUNGEN_SHEET_NAME)
        field_names = [field.name for field in fields(Schulung)]

        for column_index, field_name in enumerate(field_names, XLSX.FIRST_INDEX):
            column = get_column_letter(column_index)
            sheet[column + "1"] = field_name

            for schulung, row in XLSX.as_rows(schulungen).items():
                sheet[column + row] = getattr(schulung, field_name)
    
    @staticmethod
    def _add_scores(
            schulungen: list[Schulung],
            juleis: Iterable[JuLei],
            scores: dict[Schulung, dict[JuLei, int]],
            xlsx: Workbook
        ):
        sheet = xlsx.create_sheet(XLSX.SCORES_SHEET_NAME)

        for julei, column in XLSX.as_columns(juleis).items():
            for schulung, row in XLSX.as_rows(schulungen).items():
                sheet[column + row] = scores[schulung][julei]
