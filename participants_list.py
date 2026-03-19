
from collections import defaultdict
from collections.abc import Generator, Iterable
from dataclasses import dataclass, replace
from functools import partial
from os import makedirs
from pathlib import Path
from time import time

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from complete_data import CompleteData, XLSXWriter
from input_data import Schulung, JuLei
from xlsx_config import XLSX, Colors

@dataclass(frozen=True)
class ParticipantsLists:
    participants: dict[Schulung, list[JuLei]]
    unchecked_wishes: dict[JuLei, list[Schulung]]
    scores: dict[Schulung, dict[JuLei, int]]
    valid: bool

    @property
    def schulungen(self) -> list[Schulung]:
        schulungen = list(self.scores.keys())
        schulungen.sort(key=lambda s: s.capacity)  # optional
        return schulungen

    @property
    def juleis(self) -> list[JuLei]:
        juleis = list(next(iter(self.scores.values())).keys())
        juleis.sort(key=lambda j: (not j.from_bw, len(j.wishes)))  # optional
        return juleis

    @property
    def assigned_juleis(self) -> set[JuLei]:
        assigned_juleis: set[JuLei] = set()
        for juleis in self.participants.values():
            for julei in juleis:
                assigned_juleis.add(julei)
        return assigned_juleis

    @property
    def all_juleis_processed(self) -> bool:
        for julei, wishes in self.unchecked_wishes.items():
            if julei not in self.assigned_juleis and len(wishes) > 0:
                return False
        return True
     

def get_empty_participants_lists(data: CompleteData) -> ParticipantsLists:
    return ParticipantsLists(
        defaultdict(list),
        data.get_wishes_of_juleis(),
        data.scores,
        True
    )

def assign_juleis_ignoring_schulungs_capacity(data: ParticipantsLists) -> ParticipantsLists:
    participants = data.participants
    unchecked_wishes = data.unchecked_wishes

    for julei, wishes in unchecked_wishes.items():
        if julei in data.assigned_juleis or len(wishes) == 0:
            continue
        desired_schulung = unchecked_wishes[julei].pop(0)
        participants[desired_schulung].append(julei)

    return ParticipantsLists(participants, unchecked_wishes, data.scores, False)

def enforce_schulungs_capacity(data: ParticipantsLists) -> ParticipantsLists:
    participants = data.participants

    for schulung in participants.keys():
        if len(participants[schulung]) == schulung.capacity:
            continue
        # Lower properties are only considered, if the ones before are equal.
        participants[schulung].sort(key=lambda julei:(
            julei.from_bw,
            data.scores[schulung][julei]
        ))
        participants[schulung] = participants[schulung][:schulung.capacity]

    return ParticipantsLists(participants, data.unchecked_wishes, data.scores, True)

def get_participants_lists(data: CompleteData) -> ParticipantsLists:
    pls = get_empty_participants_lists(data)

    while not(pls.all_juleis_processed):
        pls = assign_juleis_ignoring_schulungs_capacity(pls)
        pls = enforce_schulungs_capacity(pls)
    
    return pls

class XLSXDrawer:
    COLUMN_WIDTH = 2.5
    @staticmethod
    def save_as_xlsx(data: CompleteData, output_directory: Path):
        pls = get_empty_participants_lists(data)
        rows = XLSX.as_rows(pls.schulungen)
        columns = XLSX.as_columns(pls.juleis)

        xlsx = XLSX.get_new_workbook(data.name)

        start_time = time()
        initial_sheet: Worksheet = xlsx.create_sheet(f"{time() - start_time:.6f}")
        XLSXDrawer._draw_initial_sheet(rows, columns, initial_sheet, data.get_wishes_of_juleis())

        sheet_generator = XLSXDrawer._get_sheet_generator(xlsx, initial_sheet, start_time)
        draw_participants = partial(
            XLSXDrawer._draw_participants, start_time, rows, columns, sheet_generator
        )

        while not(pls.all_juleis_processed):
            original_pls = replace(pls)
            pls = assign_juleis_ignoring_schulungs_capacity(pls)
            invalid_pls = replace(pls)
            pls = enforce_schulungs_capacity(pls)
            draw_participants(original_pls, invalid_pls, pls)

        xlsx = XLSXWriter.add_minimal_data(xlsx, data)

        makedirs(output_directory, exist_ok=True)
        xlsx.save(output_directory/f"{data.name}.xlsx")

    @staticmethod
    def _draw_initial_sheet(
            rows: dict[Schulung, str],
            columns: dict[JuLei, str],
            initial_sheet: Worksheet,
            wishes: dict[JuLei, list[Schulung]]
        ):
        for column in ["A"] + list(columns.values()):
            initial_sheet.column_dimensions[column].width = XLSXDrawer.COLUMN_WIDTH
        XLSXDrawer._draw_headers(rows, columns, initial_sheet)
        XLSXDrawer._draw_wishes(rows, columns, initial_sheet, wishes)

    @staticmethod
    def _draw_headers(rows: dict[Schulung, str], columns: dict[JuLei, str], sheet: Worksheet):
        sheet["A1"].fill = PatternFill("solid", Colors.SCHULUNG_HEADER)

        for schulung, row in rows.items():
            sheet["A"+row] = schulung.capacity
            sheet["A"+row].font = Font(bold=True, color=Colors.WHITE)
            sheet["A"+row].fill = PatternFill("solid", Colors.SCHULUNG_HEADER)
            sheet["A"+row].alignment = Alignment(horizontal='center')

        for julei, column in columns.items():
            sheet[column+"1"].font = Font(bold=True, color=Colors.WHITE)
            if julei.from_bw:
               sheet[column+"1"] = XLSX.FROM_BW_STRING
               sheet[column+"1"].fill = PatternFill("solid", Colors.JULEI_FROM_BW_HEADER)
            else:
               sheet[column+"1"] = XLSX.NOT_FROM_BW_STRING
               sheet[column+"1"].fill = PatternFill("solid", Colors.JULEI_NOT_FROM_BW_HEADER)
    
    @staticmethod
    def _draw_wishes(
            rows: dict[Schulung, str],
            columns: dict[JuLei, str],
            sheet: Worksheet,
            wishes: dict[JuLei, list[Schulung]]
        ):
        for julei, column in columns.items():
            for priority, row in enumerate(rows[s] for s in wishes[julei]):
                sheet[column+row] = priority + 1
                sheet[column+row].font = Font(color=Colors.WHITE)
                sheet[column+row].alignment = Alignment(horizontal='center')
                sheet[column+row].fill = PatternFill("solid", Colors.wish(priority, julei.from_bw))

    @staticmethod
    def _draw_participants(
            start_time: float,
            rows: dict[Schulung, str],
            columns: dict[JuLei, str],
            sheet_generator: Generator[Worksheet],
            old_data: ParticipantsLists,
            invalid_data: ParticipantsLists,
            new_data: ParticipantsLists,
        ):
        ...
ghg
    @staticmethod
    def _color_cells(
            rows: dict[Schulung, str],
            columns: dict[JuLei, str],
            sheet: Worksheet,
            schulungen: Iterable[Schulung],
            juleis: Iterable[JuLei],
            color: str
        ):
        for row in (rows[s] for s in schulungen):
            for column in (columns[j] for j in juleis):
                sheet[column+row].fill = PatternFill("solid", color)
    
    @staticmethod
    def _get_sheet_generator(
            xlsx: Workbook,
            sheet_for_duplication: Worksheet,
            start_time: float
        ) -> Generator[Worksheet]:
        sheet = sheet_for_duplication
        while True:
            sheet = xlsx.copy_worksheet(sheet)
            sheet.title = f"{time() - start_time:.6f}"
            yield sheet